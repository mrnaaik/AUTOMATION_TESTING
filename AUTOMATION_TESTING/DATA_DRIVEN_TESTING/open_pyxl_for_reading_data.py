import openpyxl
#write a code to locate an excel sheet in which data is stored
file="C:/Users/caldata.xlsx"
#write code to locate workbook
workbook=openpyxl.load_workbook(file)
#write code to locate sheet in which actual data is located
sheet=workbook["Sheet1"]
#find out how many rows has actual data
no_of_rows=sheet.max_row
print("no.of rows in table:",no_of_rows)
#find out how many columns has actual data
no_of_col=sheet.max_column
print('no.of columns from table:',no_of_col)
#print all the rows and columns from the excel sheet
#nested loops
for row in range(1,no_of_rows+1):
    for col in range(1,no_of_col+1):
        print(sheet.cell(row,col).value,end="         ")
    print("\n")
