import openpyxl
#locate an excell file without having any data
file="C:/users/writeopenpyxl.xlsx"
#locate workbook
workbook=openpyxl.load_workbook(file)
#locate sheet in which u want to write data into excell file
sheet=workbook["Sheet1"]
#write a record as matrix in each cell
#BELOW data has been written to the 1st row
sheet.cell(1,1).value=501
sheet.cell(1,2).value="anu"
sheet.cell(1,3).value="QA"
sheet.cell(1,4).value="HYD"
#below data has been written to the 2nd row
sheet.cell(2,1).value=502
sheet.cell(2,2).value="Balu"
sheet.cell(2,3).value="AIML"
sheet.cell(2,4).value="Pune"
#below data is belonging to 3rd row
sheet.cell(3,1).value=503
sheet.cell(3,2).value="Chinnu"
sheet.cell(3,3).value="Embedded system"
sheet.cell(3,4).value="HYD"
sheet.cell(4,1).value=504
sheet.cell(4,2).value="Junnu"
sheet.cell(4,3).value="QA"
sheet.cell(4,4).value="HYD"


workbook.save(file)